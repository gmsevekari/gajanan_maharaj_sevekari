import json
import re
import sys
import os

try:
    import openpyxl
except ImportError:
    print("Error: The 'openpyxl' module is required to read Excel files.")
    print("Please install it by running: pip install openpyxl")
    sys.exit(1)


def sanitize_phone(phone_raw):
    phone_str = str(phone_raw).strip()

    # Handle empty or null values
    if phone_str.lower() in ("none", "nan", ""):
        return ""

    # Extract only digits and leading plus sign if present
    cleaned = re.sub(r"[^\d+]", "", phone_str)

    if not cleaned:
        return ""

    # Ensure it starts with a '+' symbol
    if cleaned[0] != "+":
        cleaned = "+" + cleaned

    return cleaned


def parse_excel():
    file_path = "Input.xlsx"
    output_path = "preallocated_participants_1.json"

    print(f"Reading Excel file: {file_path}...")
    try:
        print("Loading workbook into memory (this may take a moment for large files)...")
        wb = openpyxl.load_workbook(file_path, data_only=True)
        print("Workbook loaded successfully. Getting active sheet...")
        sheet = wb.active
        print(f"Active sheet obtained. Reported max rows: {sheet.max_row}")
    except Exception as e:
        print(f"Failed to load workbook '{file_path}': {e}")
        sys.exit(1)

    participants = []
    failed_records = []

    # Iterating through all rows. We assume row 1 contains headers:
    # Col 1: Serial #, Col 2: Name, Col 3: Adhyays, Col 4: Phone

    empty_row_count = 0

    print("Starting row extraction loop...")

    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_idx % 50 == 0:
            print(f"Processing row {row_idx} (Extracted {len(participants)} valid participants so far)...")
            
        if row_idx == 1:
            # Skip the header row
            continue

        serial_raw = row[0]
        name_raw = row[1]
        adhyays_raw = row[2]
        phone_raw = row[3]

        # Check for completely empty rows
        if serial_raw is None and name_raw is None and phone_raw is None:
            empty_row_count += 1
            if empty_row_count > 10:
                # We've likely reached the end of the actual data
                break
            continue

        empty_row_count = 0  # Reset on valid data row

        try:
            # 1. Sanitize & Validate Serial Number
            try:
                serial = int(float(serial_raw))
            except (ValueError, TypeError):
                failed_records.append(
                    {
                        "serial": serial_raw,
                        "row": row_idx,
                        "reason": "Invalid or missing serial number.",
                    }
                )
                continue

            if not (1 <= serial <= 168):
                failed_records.append(
                    {
                        "serial": serial,
                        "row": row_idx,
                        "reason": "Serial out of range. Must be 1 to 168.",
                    }
                )
                continue

            # 2. Sanitize Name
            name = str(name_raw).strip() if name_raw is not None else ""
            if not name or name.lower() in ("none", "nan"):
                failed_records.append(
                    {"serial": serial, "row": row_idx, "reason": "Missing name."}
                )
                continue

            # 3. Sanitize Adhyays
            adhyays_str = str(adhyays_raw).strip() if adhyays_raw is not None else ""
            assigned_adhyays = []

            if adhyays_str and adhyays_str.lower() not in ("none", "nan", ""):
                valid = True
                parts = adhyays_str.split(",")
                for part in parts:
                    clean_part = part.strip()
                    if clean_part:
                        try:
                            val = int(float(clean_part))
                            if 1 <= val <= 21:
                                assigned_adhyays.append(val)
                            else:
                                valid = False
                        except ValueError:
                            valid = False

                if not valid or not assigned_adhyays:
                    failed_records.append(
                        {
                            "serial": serial,
                            "row": row_idx,
                            "reason": f"Invalid adhyays format: {adhyays_str}",
                        }
                    )
                    continue
            else:
                failed_records.append(
                    {
                        "serial": serial,
                        "row": row_idx,
                        "reason": "Missing assigned adhyays.",
                    }
                )
                continue

            # 4. Sanitize Phone
            phone = sanitize_phone(phone_raw)

            # Minimum length of 11 handles SG numbers (+65 followed by 8 digits)
            # Standard numbers like +1<10_digits> are 12 chars, +91<10_digits> are 13 chars
            if len(phone) < 11:
                failed_records.append(
                    {
                        "serial": serial,
                        "row": row_idx,
                        "reason": f"Invalid phone format: '{phone_raw}'. Resulted in '{phone}'.",
                    }
                )
                continue

            # Record matched all criteria
            participants.append(
                {
                    "index": serial,
                    "name": name,
                    "adhyays": assigned_adhyays,
                    "phone": phone,
                }
            )

        except Exception as e:
            failed_records.append(
                {
                    "serial": serial_raw,
                    "row": row_idx,
                    "reason": f"Unexpected error: {str(e)}",
                }
            )

    # Write to target JSON
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(participants, f, indent=4, ensure_ascii=False)

    print(
        f"\n✅ Successfully extracted {len(participants)} participants and written to '{output_path}'"
    )

    if failed_records:
        print("\n⚠️ Failed to extract the following Serial #s:")
        print("-" * 50)
        for fr in failed_records:
            print(f"Row {fr['row']} | Serial: {fr['serial']} | Reason: {fr['reason']}")
        print("-" * 50)
        print(f"Total failed records: {len(failed_records)}")


if __name__ == "__main__":
    parse_excel()
